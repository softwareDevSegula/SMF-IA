# traceability_engine.py — EXACT MATCHING + AUTOMATIC MULTIPLE EMISSIONS
import pandas as pd
import mariadb
import yaml
from pathlib import Path
import os

print("TN-MBSE 2025 — EXACT MATCHING + AUTOMATIC MULTIPLE EMISSIONS")
print("="*80)

# === CONFIG ===
SPEC_EXCEL = Path("../data/output_diagrams/Spec_translated.xlsx")
DB_CONFIG = Path("../01-mariadb-setup/config/database.yaml")
OUTPUT_FILE = Path("../data/output_diagrams/FINAL_TRACEABILITY_REPORT.xlsx")

# Load spec flows
if not SPEC_EXCEL.exists():
    print(f"ERROR: Spec file not found: {SPEC_EXCEL}")
    exit()

df_spec = pd.read_excel(SPEC_EXCEL)
print(f"Loaded {len(df_spec)} flows from spec")

# Load DB config
with open(DB_CONFIG) as f:
    db_cfg = yaml.safe_load(f)

conn = mariadb.connect(**db_cfg)
cur = conn.cursor()

# Get all subsystems
cur.execute("SELECT id, name FROM Subsystems ORDER BY name")
subsystems = {name: sid for sid, name in cur.fetchall()}
print(f"Found subsystems: {list(subsystems.keys())}")

# User selects primary subsystem
print("\nSELECT PRIMARY SUBSYSTEM FOR SEARCH:")
for i, name in enumerate(subsystems.keys(), 1):
    print(f"  {i}. {name}")
choice = int(input("\nEnter number: ")) - 1
primary_ss_name = list(subsystems.keys())[choice]
primary_ss_id = subsystems[primary_ss_name]

print(f"\nPRIMARY SUBSYSTEM SELECTED → {primary_ss_name}")
print("Search logic:")
print("  - CONSUMED: Find emitters in primary SS → if multiple, ask user")
print("  - EMITTED: Find ALL consumers automatically (multiple connections OK)")
print("  - ⚠️  USING EXACT FLUX NAME MATCHING ONLY")
print()

# Prepare results
results = []
ambiguous_cases = []

# Helper: find emitters of a flow (with EXACT matching)
def find_emitters(flow_name, spec_fct):
    sql = """
    SELECT s.name, f.fct_tag, fx.name
    FROM FluxEmissions fe
    JOIN Functions f ON fe.emitter_func_id = f.id
    JOIN Subsystems s ON f.subsystem_id = s.id
    JOIN Fluxes fx ON fe.flux_id = fx.id
    WHERE fx.name = ?
    """
    cur.execute(sql, (flow_name,))
    matches = cur.fetchall()
    
    if not matches:
        return None
    
    # Priority 1: Emitters in primary subsystem
    primary_ss_matches = [m for m in matches if m[0] == primary_ss_name]
    
    if len(primary_ss_matches) == 1:
        return primary_ss_matches[0]
    elif len(primary_ss_matches) > 1:
        return ("AMBIGUOUS_PRIMARY", primary_ss_matches, f"FCT_{spec_fct} in {primary_ss_name} consumes FLUX {flow_name} FROM:")
    
    # Priority 2: Emitters anywhere else
    other_matches = [m for m in matches if m[0] != primary_ss_name]
    
    if len(other_matches) == 1:
        return other_matches[0] + (f"(outside {primary_ss_name})",)
    elif len(other_matches) > 1:
        return ("AMBIGUOUS_OTHER", other_matches, f"FCT_{spec_fct} in {primary_ss_name} consumes FLUX {flow_name} FROM:")
    else:
        return None

# Helper: find ALL consumers of a flow (with EXACT matching) - AUTOMATIC MULTIPLE
def find_all_consumers(flow_name, spec_fct):
    sql = """
    SELECT s.name, f.fct_tag, fx.name
    FROM FluxConsumptions fc
    JOIN Functions f ON fc.consumer_func_id = f.id
    JOIN Subsystems s ON f.subsystem_id = s.id
    JOIN Fluxes fx ON fc.flux_id = fx.id
    WHERE fx.name = ?
    """
    cur.execute(sql, (flow_name,))
    matches = cur.fetchall()
    
    if not matches:
        return None
    
    # Filter out SD_ functions and subsystem-named functions
    valid_matches = []
    for match in matches:
        ss_name, fct_tag, flux_name = match
        
        # Skip SD_ tagged functions
        if fct_tag.startswith('SD_'):
            continue
            
        # Skip functions named exactly like the subsystem
        if fct_tag.upper() == primary_ss_name.upper():
            continue
            
        valid_matches.append(match)
    
    return valid_matches

# Main traceability logic - FIRST PASS: Collect all data
print("\n" + "="*80)
print("PHASE 1: ANALYZING FLOWS AND COLLECTING AMBIGUOUS CASES")
print("="*80)

for idx, row in df_spec.iterrows():
    spec_fct = row["Function"]
    flow = row["Flow Title"]
    direction = row["Direction"]

    if direction == "CONSUMPTION":
        # FCT consumes flux → find emitter (EXACT MATCH + USER CHOICE IF MULTIPLE)
        match = find_emitters(flow, spec_fct)
        
        if match is None:
            results.append({
                "Spec File": row["Spec File"],
                "FCT": spec_fct,
                "Flow": flow,
                "Direction": direction,
                "Status": "MISSING",
                "Found In Subsystem": "",
                "Found In FCT": "",
                "Comment": f"No emitter found for consumed flux"
            })
        elif match[0] == "AMBIGUOUS_PRIMARY":
            # Store placeholder - will be resolved by user
            result_idx = len(results)
            results.append({
                "Spec File": row["Spec File"],
                "FCT": spec_fct,
                "Flow": flow,
                "Direction": direction,
                "Status": "PENDING_USER_CHOICE",
                "Found In Subsystem": "",
                "Found In FCT": "",
                "Comment": f"Awaiting user selection from {len(match[1])} options"
            })
            ambiguous_cases.append({
                "result_idx": result_idx,
                "spec_file": row["Spec File"],
                "spec_fct": spec_fct,
                "flow": flow,
                "question": match[2],
                "options": match[1],
                "type": "CONSUMPTION"
            })
        elif match[0] == "AMBIGUOUS_OTHER":
            # Store placeholder - will be resolved by user
            result_idx = len(results)
            results.append({
                "Spec File": row["Spec File"],
                "FCT": spec_fct,
                "Flow": flow,
                "Direction": direction,
                "Status": "PENDING_USER_CHOICE",
                "Found In Subsystem": "",
                "Found In FCT": "",
                "Comment": f"Awaiting user selection from {len(match[1])} options"
            })
            ambiguous_cases.append({
                "result_idx": result_idx,
                "spec_file": row["Spec File"],
                "spec_fct": spec_fct,
                "flow": flow,
                "question": match[2],
                "options": match[1],
                "type": "CONSUMPTION"
            })
        else:
            results.append({
                "Spec File": row["Spec File"],
                "FCT": spec_fct,
                "Flow": flow,
                "Direction": direction,
                "Status": "FOUND",
                "Found In Subsystem": match[0],
                "Found In FCT": match[1],
                "Comment": match[3] if len(match) > 3 else ""
            })

    else:  # EMISSION
        # FCT emits flux → find ALL consumers (AUTOMATIC MULTIPLE CONNECTIONS)
        matches = find_all_consumers(flow, spec_fct)
        
        if not matches:
            results.append({
                "Spec File": row["Spec File"],
                "FCT": spec_fct,
                "Flow": flow,
                "Direction": direction,
                "Status": "MISSING",
                "Found In Subsystem": "",
                "Found In FCT": "",
                "Comment": "No valid consumer found"
            })
        else:
            # Create one row for EACH consumer connection (AUTOMATIC)
            for match in matches:
                ss_name, fct_tag, flux_name = match
                location = "primary SS" if ss_name == primary_ss_name else f"outside {primary_ss_name}"
                
                results.append({
                    "Spec File": row["Spec File"],
                    "FCT": spec_fct,
                    "Flow": flow,
                    "Direction": direction,
                    "Status": "FOUND",
                    "Found In Subsystem": ss_name,
                    "Found In FCT": fct_tag,
                    "Comment": f"Consumer in {location}"
                })

# === PHASE 2: USER RESOLVES AMBIGUOUS CONSUMPTION CASES ===
if ambiguous_cases:
    print(f"\n{'='*80}")
    print(f"PHASE 2: RESOLVING {len(ambiguous_cases)} AMBIGUOUS CONSUMPTION CASES")
    print(f"(Emissions are handled automatically with multiple connections)")
    print(f"{'='*80}")
    
    for i, case in enumerate(ambiguous_cases):
        print(f"\n{'─'*60}")
        print(f"CASE {i+1}/{len(ambiguous_cases)}")
        print(f"{'─'*60}")
        print(f"Spec Function: FCT_{case['spec_fct']}")
        print(f"Consuming Flux: {case['flow']}")
        print(f"\n{case['question']}")
        print(f"\nAvailable emitters:")
        
        for j, opt in enumerate(case["options"], 1):
            ss, fct, flux = opt
            print(f"  {j}. Function: {fct}")
            print(f"     Subsystem: {ss}")
            print(f"     Emits: {flux}")
            print()
        
        print(f"  {len(case['options'])+1}. ❌ MARK AS MISSING (no correct option)")
        
        while True:
            try:
                choice = int(input(f"\nYour choice (1-{len(case['options'])+1}): "))
                if 1 <= choice <= len(case['options'])+1:
                    break
                print(f"Please enter a number between 1 and {len(case['options'])+1}")
            except ValueError:
                print("Please enter a valid number")
        
        result_idx = case['result_idx']
        
        if choice <= len(case['options']):
            chosen = case['options'][choice-1]
            results[result_idx]["Found In Subsystem"] = chosen[0]
            results[result_idx]["Found In FCT"] = chosen[1]
            results[result_idx]["Status"] = "FOUND"
            results[result_idx]["Comment"] = f"User selected from {len(case['options'])} options"
            print(f"✅ Selected: {chosen[1]} in {chosen[0]}")
        else:
            results[result_idx]["Status"] = "MISSING"
            results[result_idx]["Found In Subsystem"] = ""
            results[result_idx]["Found In FCT"] = ""
            results[result_idx]["Comment"] = "User marked as missing - no correct emitter found"
            print(f"❌ Marked as MISSING")

# Final DataFrame and export
df_final = pd.DataFrame(results)

# Color coding
def color_status(val):
    if "FOUND" in val:
        return 'background-color: #d4edda'  # green
    elif "MISSING" in val:
        return 'background-color: #f8d7da'  # red
    else:
        return 'background-color: #fff3cd'  # amber

# Save with styling
with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
    df_final.to_excel(writer, index=False, sheet_name='Traceability')
    
    # Get the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Traceability']
    
    # Apply styling
    from openpyxl.styles import PatternFill
    
    green_fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
    red_fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
    
    # Find Status column (should be column E, index 5)
    status_col_idx = list(df_final.columns).index('Status') + 1
    
    for row_idx, row in enumerate(df_final.itertuples(), start=2):
        if 'FOUND' in row.Status:
            worksheet.cell(row=row_idx, column=status_col_idx).fill = green_fill
        elif 'MISSING' in row.Status:
            worksheet.cell(row=row_idx, column=status_col_idx).fill = red_fill
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width

# Summary statistics
total_connections = len(df_final)
found_connections = len(df_final[df_final['Status'].str.contains('FOUND')])
missing_connections = len(df_final[df_final['Status'].str.contains('MISSING')])
unique_flows_checked = df_spec['Flow Title'].nunique()

print("\n" + "="*80)
print("FINAL TRACEABILITY REPORT COMPLETE")
print("="*80)
print(f"   Primary Subsystem      : {primary_ss_name}")
print(f"   Unique flows checked   : {unique_flows_checked}")
print(f"   Total connections      : {total_connections}")
print(f"   ✅ Found connections    : {found_connections} ({100*found_connections/total_connections:.1f}%)")
print(f"   ❌ Missing connections  : {missing_connections} ({100*missing_connections/total_connections:.1f}%)")
print(f"\n   REPORT SAVED → {OUTPUT_FILE.resolve()}")
print("="*80)
print("PROCESSING LOGIC:")
print("   ✅ CONSUMED: Exact matching + user choice if multiple emitters")
print("   ✅ EMITTED: Exact matching + automatic multiple consumers")
print("   ✅ Filtered out SD_ and subsystem-named functions")
print("   ✅ User resolved all ambiguous consumption cases interactively")
print("\n📊 OPEN THE EXCEL FILE TO REVIEW TRACEABILITY MATRIX")
print("="*80)